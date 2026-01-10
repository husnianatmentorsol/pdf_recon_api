import re
import os
import pandas as pd
import numpy as np
import pdfplumber
from pdf2image import convert_from_bytes, convert_from_path
import pytesseract
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from googleapiclient.errors import HttpError # Added HttpError for specific error handling
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request # Added for refreshing token
from google.oauth2 import service_account # Import for service account authentication

# Configure Django settings before importing REST framework components
# This is necessary when running Django/DRF code outside a full Django project context
from django.conf import settings

# Only configure settings if not already configured
if not settings.configured:
    settings.configure(
        SECRET_KEY='a_dummy_secret_key_for_colab_testing',
        INSTALLED_APPS=[
            'django.contrib.auth',
            'django.contrib.contenttypes',
            'rest_framework',
        ],
        # Define a basic MEDIA_ROOT for file uploads
        MEDIA_ROOT=os.path.join(os.getcwd(), 'media'),
        MEDIA_URL='/media/',
        # Placeholder for Google Service Account File path
        # IMPORTANT: Replace 'path/to/your/service_account_key.json' with the actual path to your service account JSON key file
        GOOGLE_SERVICE_ACCOUNT_FILE=os.path.join(os.getcwd(), 'service_account_key.json'),
        GOOGLE_SHEET_ID = "1Z_ZKrKohFPQA_J4OKGviPBtLl7FexyKQuSbq-Hsa8JQ"
    )
    # Ensure REST_FRAMEWORK settings are configured if needed by DRF internals
    # (though defaults are usually fine for basic usage)
    settings.REST_FRAMEWORK = {
        'DEFAULT_PARSER_CLASSES': [
            'rest_framework.parsers.MultiPartParser',
            'rest_framework.parsers.FormParser',
        ]
    }

from .google_sheets_utils import get_sheets_service, create_new_tab_only

def some_view(request):
    # Sheets service create karo
    sheets_service = get_sheets_service()

    # New tab create karo
    new_tab_name = create_new_tab_only(sheets_service)

    return Response({"status": "success", "new_tab": new_tab_name})

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from .models import ReconciliationRecord
from django.utils import timezone


def save_uploaded_files(bank_file, hotel_file):
    """
    Save uploaded files to Django MEDIA folder and return their paths
    """
    upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
    os.makedirs(upload_dir, exist_ok=True)

    bank_path = os.path.join(upload_dir, bank_file.name)
    hotel_path = os.path.join(upload_dir, hotel_file.name)

    # Save bank file
    with open(bank_path, "wb+") as f:
        for chunk in bank_file.chunks():
            f.write(chunk)

    # Save hotel file
    with open(hotel_path, "wb+") as f:
        for chunk in hotel_file.chunks():
            f.write(chunk)

    return bank_path, hotel_path


def safe_float(x):
    try:
        return float(str(x).replace(",", "").strip())
    except:
        return 0.0

def extract_text_lines(pdf):
    lines=[]
    try:
        with pdfplumber.open(pdf) as p:
            for page in p.pages:
                txt = page.extract_text()
                if txt: lines += txt.split("\n")
    except pdfplumber.PDFSyntaxError:
        pass

    if not lines:
        try:
            for img in convert_from_path(pdf):
                lines += pytesseract.image_to_string(img).split("\n")
        except Exception as e:
            print(f"Error during OCR for {pdf}: {e}")

    return [l.strip() for l in lines if l.strip()]

def extract_dt(line):
    d = re.search(r"\d{2}-[A-Za-z]{3}-\d{4}", line)
    t = re.search(r"\d{2}:\d{2}(?:\u2192\u0102\u0100\u0102\u0100\u0102\u0100\d{2})?", line)
    date = d.group() if d else ""
    time = t.group() if t else "00:00:00"
    if len(time)==5: time += ":00"
    return date, time

def add_titles_and_total(df, titles, column_headers, amount_cols_to_sum=None):
    sheet_data = []
    for t in titles:
        sheet_data.append([t])
    sheet_data.append([])

    sheet_data.append(column_headers)

    if df.empty:
        empty_row = ["No Records"] + [''] * (len(column_headers) - 1)
        sheet_data.append(empty_row)
    else:
        df_to_export = df.copy()

        if amount_cols_to_sum and all(col in df_to_export.columns for col in amount_cols_to_sum):
            total_row_dict = {col: '' for col in df_to_export.columns}
            for amount_col in amount_cols_to_sum:
                total_row_dict[amount_col] = df_to_export[amount_col].sum()

            if amount_cols_to_sum:
                first_amount_col_idx = df_to_export.columns.get_loc(amount_cols_to_sum[0])
                if first_amount_col_idx > 0:
                    col_before_first_amount = df_to_export.columns[first_amount_col_idx - 1]
                    total_row_dict[col_before_first_amount] = "TOTAL"
                else:
                    total_row_dict[df_to_export.columns[0]] = "TOTAL"

            total_row_df = pd.DataFrame([total_row_dict], columns=df_to_export.columns)
            df_to_export = pd.concat([df_to_export, total_row_df], ignore_index=True)

        sheet_data.extend(df_to_export.values.tolist())

    return pd.DataFrame(sheet_data)


def bank_df(lines):
    rows=[]
    merchant_id = ""
    terminal_id = ""

    for l in lines[:20]:
        mid_match = re.search(r"ID (\S+)", l)
        if mid_match:
            merchant_id = mid_match.group(1)
        tid_match = re.search(r"TERMINAL ID (\S+)", l)
        if tid_match:
            terminal_id = tid_match.group(1)

    bank_pattern = re.compile(
        r"^\d+\s+"
        r"(\d{2}/\d{2}/\d{4})\s+"
        r"(\d{2}:\d{2})\s+"
        r"(?:\d{2})\s+"
        r"(\S+)\s+"
        r"(\S+)\s+"
        r"([\d,]+\.\d{2})\s+"
        r"([\d,]+\.\d{2})\s+"
        r"([\d,]+\.\d{2})"
        r"$"
    )

    card_type_header_pattern = re.compile(r"(ON-US|OFF-US)\s+(VISA|MASTERCARD|NAPS|GCCNET|AMEX|DINERS|JCB)", re.IGNORECASE)

    GCCNET_CARD_LAST_4_DIGITS = {"0580", "8628", "8134"}

    current_card_type = "UNKNOWN"

    for l in lines:
        card_header_match = card_type_header_pattern.search(l)
        if card_header_match:
            current_card_type = card_header_match.group(2).upper()
            continue

        match = bank_pattern.match(l)
        if match:
            (date_str, time_str, ref_num, card_num,
             gross_amount_str, commission_str, net_amount_str) = match.groups()

            gross_amount = safe_float(gross_amount_str)
            commission = safe_float(commission_str)
            net_amount = safe_float(net_amount_str)

            card_type_display = current_card_type

            if card_num[-4:] in GCCNET_CARD_LAST_4_DIGITS:
                card_type_display = 'GCCNET'

            rows.append([
                date_str, time_str,
                merchant_id, ref_num, card_num, card_type_display,
                gross_amount, commission, net_amount, terminal_id
            ])
    df=pd.DataFrame(rows,columns=[
        "Transaction Date","Time","Merchant ID","Invoice No / RRN",
        "Card Number","Card Type (On us/Off us)","Gross Amount","Commission","Net Amount","Terminal ID"
    ])
    df["DT"]=pd.to_datetime(df["Transaction Date"].str.replace('/','-')+" "+df["Time"], format="%d-%m-%Y %H:%M", errors="coerce")
    return df

# ===============================
# HOTEL PDF -> VISA SETTLEMENTS (Attachment 6)
# ===============================
def hotel_df(lines):
    rows=[]
    unmatched_lines = []

    hotel_txn_pattern = re.compile(
        r"^"
        r"(\d{2}/\d{2}/\d{2})\s+"
        r"(\d{2}:\d{2})\s+"
        r"(\S+)\s+"
        r"(.*?)\s*"
        r"(?P<txn_code>\d{5})\s+"
        r"((?:POS - )?(?:Visa|Master|Amex|NAPS|GCCNET|Other)(?: Card)?)\s+"
        r"(?:(\S+)\s*)?" # Corrected: Removed 'standardised' and ensured regex termination
        r"(QAR)\s+"
        r"([\d,]+\.\d{2})\s*(?:-?\s*)?"
        r"([\d,]+\.\d{2})\s*"
        r"(\S+)"
        r"$"
    )

    card_num_pattern = re.compile(r"^(\S{4}X+\d{4})")
    check_ref_pattern = re.compile(r"CHECK#\s*(\d+)\s*\[(\d+)\]")

    i = 0
    while i < len(lines):
        l = lines[i]
        match = hotel_txn_pattern.match(l)
        if match:
            (date_str, time_str, room_no, name_desc, txn_code, card_type_full, check_ref_num, currency, debit_amount_str, credit_amount_str, cashier_id) = match.groups()

            card_ref = ""
            card_ref_parts = []

            if check_ref_num:
                check_match_in_ref_num = check_ref_pattern.search(check_ref_num)
                if check_match_in_ref_num:
                    card_ref_parts.append(f"CHECK# {check_match_in_ref_num.group(1)} [{check_match_in_ref_num.group(2)}]") # Fixed typo here
                else:
                    card_ref_parts.append(check_ref_num)

            current_next_line_idx = i + 1
            if current_next_line_idx < len(lines):
                next_line_content = lines[current_next_line_idx].strip()
                check_match_next_line = check_ref_pattern.search(next_line_content)
                if check_match_next_line:
                    card_ref_parts.append(next_line_content)
                    i += 1
                    current_next_line_idx = i + 1

            if current_next_line_idx < len(lines):
                next_line_content = lines[current_next_line_idx].strip()
                card_num_match_next_line = card_num_pattern.match(next_line_content)
                if card_num_match_next_line:
                    card_ref_parts.append(card_num_match_next_line.group(1))
                    i += 1

            if card_ref_parts:
                card_ref = " / ".join(card_ref_parts)
            else:
                card_ref = ""

            rows.append([
                date_str, time_str,
                room_no, name_desc.strip(), card_ref, card_type_full,
                safe_float(credit_amount_str), cashier_id
            ])
        else:
            if l.strip() and not check_ref_pattern.match(l) and not card_num_pattern.match(l):
                unmatched_lines.append(l.strip())
        i += 1

    if unmatched_lines:
        print("\n--- Unmatched lines from HOTEL PDF (potential missing transactions) ---")
        for ul in unmatched_lines:
            print(ul)

    df=pd.DataFrame(rows,columns=[
        "Transaction Date","Time","Room No","Name","Card Reference","Card Type","Amount","Cashier ID"
    ])
    df["DT"]=pd.to_datetime(df["Transaction Date"].str.replace('/','-')+" "+df["Time"], format="%d-%m-%y %H:%M", errors="coerce")
    return df



rec_bank_list = []
rec_hotel_list = []
un_bank_list = []

BANK_COLUMNS = ["Transaction Date","Time","Merchant ID","Invoice No / RRN",
                "Card Number","Card Type (On us/Off us)","Gross Amount","Commission","Net Amount","Terminal ID"]
HOTEL_COLUMNS = ["Transaction Date","Time","Room No","Name","Card Reference","Card Type","Amount","Cashier ID"]

bank = pd.DataFrame(columns=BANK_COLUMNS)
hotel = pd.DataFrame(columns=HOTEL_COLUMNS)

un_hotel_df = pd.DataFrame(columns=HOTEL_COLUMNS)

rec_bank = pd.DataFrame(columns=BANK_COLUMNS)
rec_hotel = pd.DataFrame(columns=HOTEL_COLUMNS)
un_bank = pd.DataFrame(columns=BANK_COLUMNS)
un_hotel = pd.DataFrame(columns=HOTEL_COLUMNS)


def normalize_card_type(card_type_str):
    """
    Normalize card type strings to standard names.
    """
    if isinstance(card_type_str, str):
        normalized = card_type_str.replace(' Card', '').replace('POS - ', '').upper().strip()
        if normalized == 'MASTER':
            return 'MASTERCARD'
        return normalized
    return card_type_str

categorized_rec_bank = {}
categorized_rec_hotel = {}
categorized_un_bank = {}
categorized_un_hotel = {}
final_card_types = []


def generate_attachment_info(
    categorized_rec_bank,
    categorized_rec_hotel,
    categorized_un_bank,
    categorized_un_hotel,
    empty_bank_df,
    empty_hotel_df,
    final_card_types
):
    attachment_info = {}
    attachment_counter = 1

    preferred_card_types_order = ['VISA', 'MASTERCARD', 'NAPS', 'GCCNET']

    all_card_types_in_use = set(final_card_types) | set(preferred_card_types_order)

    final_card_types_sorted = sorted(list(all_card_types_in_use))

    for card_type in final_card_types_sorted:
        rec_bank_df = categorized_rec_bank.get(card_type, empty_bank_df.copy())
        titles_rec_bank = [
            f"Attachment - {attachment_counter}",
            f"{card_type} Merchant Transactions",
            f"Reconciled {card_type} Transactions"
        ]
        attachment_info[f"Attachment {attachment_counter}"] = (
            rec_bank_df, titles_rec_bank, BANK_COLUMNS, ["Gross Amount", "Commission", "Net Amount"]
        )
        attachment_counter += 1

        rec_hotel_df = categorized_rec_hotel.get(card_type, empty_hotel_df.copy())
        titles_rec_hotel = [
            f"Attachment - {attachment_counter}",
            f"{card_type} Settlements",
            f"Reconciled {card_type} Transactions"
        ]
        attachment_info[f"Attachment {attachment_counter}"] = (
            rec_hotel_df, titles_rec_hotel, HOTEL_COLUMNS, ["Amount"]
        )
        attachment_counter += 1

        un_bank_df_cat = categorized_un_bank.get(card_type, empty_bank_df.copy())
        titles_un_bank = [
            f"Attachment - {attachment_counter}",
            f"{card_type} Merchant Transactions",
            f"Unreconciled {card_type} Transactions"
        ]
        attachment_info[f"Attachment {attachment_counter}"] = (
            un_bank_df_cat, titles_un_bank, BANK_COLUMNS, ["Gross Amount", "Commission", "Net Amount"]
        )
        attachment_counter += 1

        un_hotel_df_cat = categorized_un_hotel.get(card_type, empty_hotel_df.copy())
        titles_un_hotel = [
            f"Attachment - {attachment_counter}",
            f"{card_type} Settlements",
            f"Unreconciled / Outstanding {card_type} Transactions"
        ]
        attachment_info[f"Attachment {attachment_counter}"] = (
            un_hotel_df_cat, titles_un_hotel, HOTEL_COLUMNS, ["Amount"]
        )
        attachment_counter += 1

    return attachment_info

summary_data_dynamic = []

def get_formatted_amount(amount):
    return f"{amount:,.2f}" if amount != 0.0 else "-"

attachment_num_lookup = {}

def save_df_to_html(dataframes_to_html, file_path, main_report_title="Credit Card Reconciliation Report"):
    html_content = []
    html_content.append("<!DOCTYPE html>")
    html_content.append("<html>")
    html_content.append("<head>")
    html_content.append(f"<title>{main_report_title}</title>")
    html_content.append("    <style>")
    html_content.append("        body { font-family: sans-serif; margin: 20px; background-color: #f4f4f4; }")
    html_content.append("        h1 { color: #333; text-align: center; margin-bottom: 30px; font-size: 2.5em; }")
    html_content.append("        h2 { color: #555; border-bottom: 2px solid #ddd; padding-bottom: 10px; margin-top: 40px; font-size: 1.8em; }")
    html_content.append("        table { width: 100%; border-collapse: collapse; margin-bottom: 30px; background-color: #fff; box-shadow: 0 2px 3px rgba(0,0,0,0.1); }")
    html_content.append("        th, td { border: 1px solid #ddd; padding: 12px 15px; text-align: left; }")
    html_content.append("        th { background-color: #e9e9e9; font-weight: bold; color: #333; text-transform: uppercase; }")
    html_content.append("        tr:nth-child(even) { background-color: #f9f9f9; }")
    html_content.append("        tr:hover { background-color: #f1f1f1; }")
    html_content.append("        .center-align { text-align: center; }")
    html_content.append("        .total-row td { background-color: #e0e0e0; font-weight: bold; }")
    html_content.append("    </style>")
    html_content.append("</head>")
    html_content.append("<body>")
    html_content.append(f"<h1>{main_report_title}</h1>")

    for title, df in dataframes_to_html:
        html_content.append(f"<h2>{title}</h2>")
        # Convert DataFrame to HTML, avoiding header=False for better styling control and adding classes
        df_html = df.to_html(index=False, header=False)

        # Replace default table tag with styled one
        df_html = df_html.replace('<table border="1" class="dataframe">', '<table>')

        # Identify the header row based on content of add_titles_and_total
        # The column headers are typically the 3rd row (index 2) in the DataFrame returned by add_titles_and_total
        # This assumes that the first dataframe in the list has this structure
        if title == "Bank Account Summary" and len(df) > 10:
            # The 10th row of the Bank Account Summary is 'Ending Balance...' headers
            # The 11th row is the actual data
            # More sophisticated header styling would go here if we were to reconstruct the HTML table more granularly
            pass

        html_content.append(df_html)

    html_content.append("</body>")
    html_content.append("</html>")

    with open(file_path, "w", encoding="utf-8") as f:
        f.write("\n".join(html_content))


class ReconciliationAPIView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        bank_file_obj = request.FILES.get("bank_file")
        hotel_file_obj = request.FILES.get("hotel_file")
        client_name = (request.data.get("client_name") or "client").strip()
        threshold_time_raw = request.data.get("threshold_time", 30)

        if not bank_file_obj or not hotel_file_obj:
            return Response({"error": "Please upload both Bank and Hotel files."},
                            status=400)
        if not client_name:
            client_name = "client"
        try:
            threshold_minutes = int(threshold_time_raw)
        except (TypeError, ValueError):
            return Response({"error": "threshold_time must be an integer number of minutes."},
                            status=400)
        if threshold_minutes < 0:
            return Response({"error": "threshold_time must be a non-negative integer."},
                            status=400)

        bank_file_path, hotel_file_path = save_uploaded_files(bank_file_obj, hotel_file_obj)

        bank_lines = extract_text_lines(bank_file_path)
        hotel_lines = extract_text_lines(hotel_file_path)

        bank = bank_df(bank_lines)
        hotel = hotel_df(hotel_lines)

        if not bank.empty:
            min_dt = bank["DT"].min()
            max_dt = bank["DT"].max()
            
            if pd.isna(min_dt) or pd.isna(max_dt):
                 pass 
            else:
                min_date = min_dt.date()
                max_date = max_dt.date()
                txn_count = len(bank)

                if ReconciliationRecord.objects.filter(
                    client_name=client_name,
                    min_date=min_date,
                    max_date=max_date,
                    total_transactions=txn_count,
                ).exists():
                     return Response(
                        {"error": f"Transactions from {min_date} to {max_date} ({txn_count} entries) have already been reconciled."},
                        status=400
                     )

        BANK_COLUMNS_DYNAMIC = bank.columns.tolist() if not bank.empty else ["Transaction Date","Time","Merchant ID","Invoice No / RRN",
                                                                           "Card Number","Card Type (On us/Off us)","Gross Amount","Commission","Net Amount","Terminal ID"]
        HOTEL_COLUMNS_DYNAMIC = hotel.columns.tolist() if not hotel.empty else ["Transaction Date","Time","Room No","Name","Card Reference","Card Type","Amount","Cashier ID"]

        empty_bank_df = pd.DataFrame(columns=BANK_COLUMNS_DYNAMIC)
        empty_hotel_df = pd.DataFrame(columns=HOTEL_COLUMNS_DYNAMIC)

        rec_bank_list = []
        rec_hotel_list = []
        un_bank_list = []
        un_hotel_df = hotel.copy() # Initialize un_hotel_df with all hotel transactions

        if not bank.empty and not un_hotel_df.empty:
            for i, b in bank.iterrows():
                match = pd.DataFrame()

                if b["Card Type (On us/Off us)"] == 'GCCNET':
                    match = un_hotel_df[
                        (un_hotel_df["Amount"] == b["Gross Amount"]) &
                        (abs(un_hotel_df["DT"] - b["DT"]) <= timedelta(minutes=threshold_minutes))
                    ]
                else:
                    bank_card_last_4 = str(b["Card Number"])[-4:]
                    potential_matches = un_hotel_df[
                        (un_hotel_df["Card Reference"].str.len() >= 4) &
                        (un_hotel_df["Card Reference"].str[-4:] == bank_card_last_4) &
                        (un_hotel_df["Amount"] == b["Gross Amount"]) &
                        (abs(un_hotel_df["DT"] - b["DT"]) <= timedelta(minutes=threshold_minutes))
                    ]
                    match = potential_matches

                    if match.empty:
                        match = un_hotel_df[
                            (un_hotel_df["Amount"] == b["Gross Amount"]) &
                            (abs(un_hotel_df["DT"] - b["DT"]) <= timedelta(minutes=threshold_minutes))
                        ]

                if not match.empty:
                    rec_bank_list.append(b)
                    rec_hotel_list.append(match.iloc[0])
                    un_hotel_df = un_hotel_df.drop(match.index[0])
                else:
                    un_bank_list.append(b)

        rec_bank = pd.DataFrame(rec_bank_list, columns=BANK_COLUMNS_DYNAMIC)
        rec_hotel = pd.DataFrame(rec_hotel_list, columns=HOTEL_COLUMNS_DYNAMIC)
        un_bank = pd.DataFrame(un_bank_list, columns=BANK_COLUMNS_DYNAMIC)
        un_hotel = un_hotel_df.copy()

        # Drop helper 'DT' column if exists
        for df_reco in [rec_bank, rec_hotel, un_bank, un_hotel]:
            df_reco.drop(columns=["DT"], errors="ignore", inplace=True)


        # ===============================
        # Step 5: Normalize Card Types & Categorize Transactions (Moved from global scope)
        # ===============================

        # Collect all card types from bank and hotel
        all_bank_card_types = pd.Series(dtype=str)
        if not rec_bank.empty:
            all_bank_card_types = pd.concat([all_bank_card_types, rec_bank['Card Type (On us/Off us)']])
        if not un_bank.empty:
            all_bank_card_types = pd.concat([all_bank_card_types, un_bank['Card Type (On us/Off us)']])

        all_hotel_card_types = pd.Series(dtype=str)
        if not rec_hotel.empty:
            all_hotel_card_types = pd.concat([all_hotel_card_types, rec_hotel['Card Type']])
        if not un_hotel.empty:
            all_hotel_card_types = pd.concat([all_hotel_card_types, un_hotel['Card Type']])

        # Apply normalization to hotel card types
        norm_hotel_card_types = all_hotel_card_types.apply(normalize_card_type)

        # Combine all card types and filter out unwanted
        unique_card_types = pd.concat([all_bank_card_types, norm_hotel_card_types]).dropna().unique()
        unique_card_types = [ct.strip() for ct in unique_card_types if ct.strip() != '']
        unique_card_types = [ct for ct in unique_card_types if ct not in ['AMEX', 'DINERS', 'JCB']]

        # Ensure mandatory card types are included
        for mandatory in ['VISA', 'MASTERCARD', 'NAPS', 'GCCNET']:
            if mandatory not in unique_card_types:
                unique_card_types.append(mandatory)

        final_card_types = sorted(unique_card_types) # Update final_card_types for this run

        # Initialize categorized DataFrames dictionaries
        categorized_rec_bank = {ct: pd.DataFrame(columns=BANK_COLUMNS_DYNAMIC) for ct in final_card_types}
        categorized_rec_hotel = {ct: pd.DataFrame(columns=HOTEL_COLUMNS_DYNAMIC) for ct in final_card_types}
        categorized_un_bank = {ct: pd.DataFrame(columns=BANK_COLUMNS_DYNAMIC) for ct in final_card_types}
        categorized_un_hotel = {ct: pd.DataFrame(columns=HOTEL_COLUMNS_DYNAMIC) for ct in final_card_types}

        # Categorize transactions by card type
        for card_type in final_card_types:
            if not rec_bank.empty:
                categorized_rec_bank[card_type] = rec_bank[rec_bank['Card Type (On us/Off us)'] == card_type].copy()
            if not un_bank.empty:
                categorized_un_bank[card_type] = un_bank[un_bank['Card Type (On us/Off us)'] == card_type].copy()

            if not rec_hotel.empty:
                categorized_rec_hotel[card_type] = rec_hotel[rec_hotel['Card Type'].apply(normalize_card_type) == card_type].copy()
            if not un_hotel.empty:
                categorized_un_hotel[card_type] = un_hotel[un_hotel['Card Type'].apply(normalize_card_type) == card_type].copy()


        # ===============================
        # Step 6: Generate attachment_info
        # ===============================
        attachment_info = generate_attachment_info(
            categorized_rec_bank,
            categorized_rec_hotel,
            categorized_un_bank,
            categorized_un_hotel,
            empty_bank_df,
            empty_hotel_df,
            final_card_types # Pass final_card_types
        )

        # ===============================
        # Step 7: Build summary_data_dynamic
        # ===============================
        summary_data_dynamic = []
        attachment_num_lookup = {}

        def get_formatted_amount(amount):
            return f"{amount:,.2f}" if amount != 0.0 else "-"

        # Map attachment numbers for quick lookup
        for att_key, (df, titles, cols, amount_cols_to_sum) in attachment_info.items():
            attachment_num = att_key.split(' ')[1]
            card_type_from_title = titles[1].split(' ')[0].upper()
            transaction_type_descriptor = 'Merchant' if 'Merchant' in titles[1] else 'Settlements'
            reconciliation_status = 'Reconciled' if 'Reconciled' in titles[2] else 'Unreconciled'
            lookup_key = (card_type_from_title, transaction_type_descriptor, reconciliation_status)
            attachment_num_lookup[lookup_key] = attachment_num

        # -------------------------------
        # Reconciled transactions
        # -------------------------------
        summary_data_dynamic.append(["Reconciled Transactions:", "", "", "", "", "Reconciled Transactions:", "", "", ""])
        for card_type in final_card_types:
            rec_bank_df = categorized_rec_bank.get(card_type, empty_bank_df)
            rec_hotel_df = categorized_rec_hotel.get(card_type, empty_hotel_df)

            rec_bank_entries = len(rec_bank_df)
            rec_bank_amount = rec_bank_df['Gross Amount'].sum() if not rec_bank_df.empty else 0.0
            att_rec_bank = attachment_num_lookup.get((card_type, 'Merchant', 'Reconciled'), '')

            rec_hotel_entries = len(rec_hotel_df)
            rec_hotel_amount = rec_hotel_df['Amount'].sum() if not rec_hotel_df.empty else 0.0
            att_rec_hotel = attachment_num_lookup.get((card_type, 'Settlements', 'Reconciled'), '')

            summary_data_dynamic.append(
                [
                    card_type, "",
                    f"Attachment {att_rec_bank}" if att_rec_bank else "-",
                    rec_bank_entries,
                    get_formatted_amount(rec_bank_amount),
                    card_type,
                    f"Attachment {att_rec_hotel}" if att_rec_hotel else "-",
                    rec_hotel_entries,
                    get_formatted_amount(rec_hotel_amount)
                ]
            )

        # -------------------------------
        # Unreconciled transactions
        # -------------------------------
        summary_data_dynamic.append(["", "", "", "", "", "", "", "", ""])
        summary_data_dynamic.append(["Credited Amounts not Recorded in Opera PMS", "", "", "", "", "Outstanding Amounts not Credited in Bank", "", "", ""])
        for card_type in final_card_types:
            un_bank_df = categorized_un_bank.get(card_type, empty_bank_df)
            un_hotel_df_cat = categorized_un_hotel.get(card_type, empty_hotel_df)

            un_bank_entries = len(un_bank_df)
            un_bank_amount = un_bank_df['Gross Amount'].sum() if not un_bank_df.empty else 0.0
            att_un_bank = attachment_num_lookup.get((card_type, 'Merchant', 'Unreconciled'), '')

            un_hotel_entries = len(un_hotel_df_cat)
            un_hotel_amount = un_hotel_df_cat['Amount'].sum() if not un_hotel_df_cat.empty else 0.0
            att_un_hotel = attachment_num_lookup.get((card_type, 'Settlements', 'Unreconciled'), '')

            summary_data_dynamic.append(
                [
                    card_type, "",
                    f"Attachment {att_un_bank}" if att_un_bank else "-",
                    un_bank_entries,
                    get_formatted_amount(un_bank_amount),
                    card_type,
                    f"Attachment {att_un_hotel}" if att_un_hotel else "-",
                    un_hotel_entries,
                    get_formatted_amount(un_hotel_amount)
                ]
            )

        # Footer rows
        summary_data_dynamic.append(["", "", "", "", "", "", "", "", ""])
        summary_data_dynamic.append(["Variance", "", "", "0", "-", "Ending Actual Net Cash Balance", "", "0", "-"])
        summary_data_dynamic.append(["", "", "", "", "", "", "", "", ""])
        summary_data_dynamic.append(["Reviewed BY", "", "", "", "", "Approved BY", "", "", ""])
        summary_data_dynamic.append(["______________", "", "", "", "", "______________", "", "", ""])

        summary_data_updated = summary_data_dynamic

        bank_ending_balance = bank['Gross Amount'].sum() if not bank.empty else 0.0
        hotel_ending_balance = hotel['Amount'].sum() if not hotel.empty else 0.0

        rows = [
            ["Company Name (Update Me)","","","","","","","",""],
            ["Credit Card Reconciliation","","","","","","","",""],
            ["","","","","","","","",""],
            ["Reconciliation Date","", datetime.now().strftime("%d-%b-%Y"),"","","","","",""],
            ["Account Name:","", "Example Hotel LLC","","","","","",""],
            ["Account Number:","", "1234-5678-9012-3456","","","","","",""],
            ["Bank Name:","", "QNB Al-Najada Branch","","","","","",""],
            ["General Ledger Account #","", "GL-C/C-4001","","","","","",""],
            ["","","","","","","","",""],
            ["Ending Balance as per Bank Statement","","Reference","Entries","Amount",
             "Ending Balance as per General Ledger","Reference","Entries","Amount"],
            ["", "", "", len(bank), f"{bank_ending_balance:,.2f}", "", "", len(hotel), f"{hotel_ending_balance:,.2f}"],
        ]
        rows.extend(summary_data_updated)

        # Convert rows to DataFrame for HTML preview
        bank_account_df = pd.DataFrame(rows)

        # Generate a timestamp for the filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_folder_name = f"{client_name}_run_{timestamp}"
        unique_folder_path = os.path.join(settings.MEDIA_ROOT, unique_folder_name)
        os.makedirs(unique_folder_path, exist_ok=True)

        report_file_name = f"Credit_Card_Reconciliation_{timestamp}.xlsx"
        path = os.path.join(unique_folder_path, report_file_name)

        html_preview_file_name = "Credit_Card_Reconciliation_Report.html"
        html_preview_file_path = os.path.join(unique_folder_path, html_preview_file_name)

        dataframes_for_html_preview = [
            ("Bank Account Summary", bank_account_df.copy())
        ]

        with pd.ExcelWriter(path, engine="openpyxl") as w:
            bank_account_df.to_excel(w, index=False, header=False, sheet_name="Bank Account")

            for name, (df, titles, cols, amount_cols_to_sum) in attachment_info.items():
                df_final = add_titles_and_total(
                    df, titles, cols, amount_cols_to_sum
                )
                df_final.to_excel(w, index=False, header=False, sheet_name=name)
                # Add attachment dataframes to the list for HTML preview
                dataframes_for_html_preview.append((f"Attachment {name.split(' ')[1]} - {titles[1]} ({titles[2]})", df_final.copy()))

        # Generate HTML preview with all sheets
        save_df_to_html(dataframes_for_html_preview, html_preview_file_path, main_report_title='Credit Card Reconciliation Report')


        wb=load_workbook(path)

        center_aligned_text = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thick_border = Side(border_style="medium", color="000000")

        ws_bank = wb["Bank Account"]
        ws_bank.merge_cells("A1:I1"); ws_bank.merge_cells("A2:I2")
        ws_bank["A1"].font=Font(bold=True,size=14)
        ws_bank["A2"].font=Font(bold=True)
        ws_bank["A1"].alignment = ws_bank["A2"].alignment = center_aligned_text

        ws_bank["A1"].fill = header_fill
        ws_bank["A2"].fill = header_fill

        for col_idx in range(1, 10):
            cell = ws_bank.cell(row=10, column=col_idx)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = center_aligned_text

        for col_idx in range(1, 10):
            cell = ws_bank.cell(row=11, column=col_idx)
            cell.fill = total_fill
            cell.font = Font(bold=True)
            cell.alignment = center_aligned_text

        for row in ws_bank.iter_rows():
            for c in row:
                c.alignment = center_aligned_text
                c.border = Border(left=thick_border,right=thick_border,top=thick_border,bottom=thick_border)

        for r_idx, row in enumerate(ws_bank.iter_rows()):
            for c_idx, c in enumerate(row):
                if isinstance(c.value,str) and c.value.startswith("Attachment"):
                    target_sheet_name = c.value.replace('Attachment - ', 'Attachment ')
                    if target_sheet_name in wb.sheetnames:
                        c.hyperlink=f"#'{target_sheet_name}'!A1"
                        c.font=Font(color="0000FF",underline="single")

        for col in range(1,10):
            ws_bank.column_dimensions[get_column_letter(col)].width = 20

        for sheet_name in attachment_info.keys():
            ws=wb[sheet_name]
            max_col = ws.max_column
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            ws["A1"].font=Font(bold=True,size=14)
            ws["A1"].alignment = center_aligned_text
            ws["A1"].fill = header_fill

            if ws.max_row >= 6:
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=5, column=col_idx)
                    cell.fill = header_fill
                    cell.font = Font(bold=True)
                    cell.alignment = center_aligned_text

            for row in ws.iter_rows():
                for c in row:
                    c.alignment = center_aligned_text
                    c.border = Border(left=thick_border,right=thick_border,top=thick_border,bottom=thick_border)

            last_row_num = ws.max_row
            if last_row_num > 6 and ws.cell(row=last_row_num, column=1).value == "TOTAL":
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=last_row_num, column=col_idx)
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
                    cell.alignment = center_aligned_text

            for col_idx in range(1, ws.max_column+1):
                ws.column_dimensions[get_column_letter(col_idx)].width=20

        for card_type in final_card_types:
            att_un_bank = attachment_num_lookup.get((card_type, 'Merchant', 'Unreconciled'), None)
            if att_un_bank:
                sheet_name_bank = f"Attachment {att_un_bank}"
                if sheet_name_bank in wb.sheetnames:
                    ws_bank_un = wb[sheet_name_bank]
                    bank_amount_cols = ["Gross Amount", "Commission", "Net Amount"]
                    for col_name in bank_amount_cols:
                        if col_name in BANK_COLUMNS_DYNAMIC:
                            col_idx = BANK_COLUMNS_DYNAMIC.index(col_name) + 1
                            for r in range(6, ws_bank_un.max_row+1):
                                cell = ws_bank_un.cell(row=r, column=col_idx)
                                if isinstance(cell.value,(int,float)):
                                    cell.number_format = '#,##0.00'
                            last_row = ws_bank_un.max_row
                            total_cell = ws_bank_un.cell(row=last_row, column=col_idx)
                            if isinstance(total_cell.value,(int,float)):
                                total_cell.font = Font(bold=True)
                                total_cell.number_format = '#,##0.00'

                    last_row = ws_bank_un.max_row
                    if last_row > 1 and ws_bank_un.cell(row=last_row, column=1).value == "TOTAL":
                        ws_bank_un.cell(row=last_row, column=1).font = Font(bold=True)

            att_un_hotel = attachment_num_lookup.get((card_type, 'Settlements', 'Unreconciled'), None)
            if att_un_hotel:
                sheet_name_hotel = f"Attachment {att_un_hotel}"
                if sheet_name_hotel in wb.sheetnames:
                    ws_hotel_un = wb[sheet_name_hotel]
                    hotel_amount_cols = ["Amount"]
                    for col_name in hotel_amount_cols:
                        if col_name in HOTEL_COLUMNS_DYNAMIC:
                            col_idx = HOTEL_COLUMNS_DYNAMIC.index(col_name) + 1
                            for r in range(6, ws_hotel_un.max_row+1):
                                cell = ws_hotel_un.cell(row=r, column=col_idx)
                                if isinstance(cell.value,(int,float)):
                                    cell.number_format = '#,##0.00'
                            last_row = ws_hotel_un.max_row
                            total_cell = ws_hotel_un.cell(row=last_row, column=col_idx)
                            if isinstance(total_cell.value,(int,float)):
                                total_cell.font = Font(bold=True)
                                total_cell.number_format = '#,##0.00'

                    last_row = ws_hotel_un.max_row
                    if "Amount" in HOTEL_COLUMNS_DYNAMIC:
                        amount_col_idx_hotel = HOTEL_COLUMNS_DYNAMIC.index("Amount")
                        if amount_col_idx_hotel > 0:
                            total_label_col_idx = amount_col_idx_hotel
                            if last_row > 1 and ws_hotel_un.cell(row=last_row, column=total_label_col_idx).value == "TOTAL":
                                ws_hotel_un.cell(row=last_row, column=total_label_col_idx).font = Font(bold=True)


                # Excel file save
        wb.save(path)

        # ===========================
        # Reconciliation Counts
                # ===========================
        # Reconciliation Counts
        # ===========================
        reconciledCount = len(rec_bank) + len(rec_hotel)
        unreconciledCount = len(un_bank) + len(un_hotel)
        totalEntries = len(bank) + len(hotel)

        # ===========================
        # Google Drive Configuration and Upload
        # =====================
        # Modify SCOPES for Google Drive
        SCOPES = ["https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/spreadsheets"]
        drive_file_preview_url = None
        drive_file_download_url = None
        google_sheet_link = None # Initialize new variable

        drive_service = None
        sheets_service = None # Initialize sheets_service here too
        creds = None # Initialize creds to None

        # --- Attempt Service Account Authentication First ---
        try:
            # Check if GOOGLE_SERVICE_ACCOUNT_FILE exists and is valid
            if os.path.exists(settings.GOOGLE_SERVICE_ACCOUNT_FILE):
                creds = service_account.Credentials.from_service_account_file(
                    settings.GOOGLE_SERVICE_ACCOUNT_FILE,
                    scopes=SCOPES
                )
                print("Google Drive and Sheets connected successfully using Service Account authentication.")
            else:
                print("GOOGLE_SERVICE_ACCOUNT_FILE not found. Attempting OAuth Client authentication.")

        except HttpError as err:
            print(f"Service Account authentication failed (HttpError): {err}")
            creds = None # Ensure creds is None if service account fails
        except Exception as e:
            print(f"Service Account authentication failed (General Error): {e}")
            creds = None # Ensure creds is None if service account fails

        # --- If Service Account failed or not configured, attempt OAuth Client (user-based) from token.json ---
        if not creds:
            token_path = os.path.join(os.getcwd(), "token.json") # Assumes token.json is in the current working directory
            if os.path.exists(token_path):
                try:
                    creds = Credentials.from_authorized_user_file(token_path, SCOPES)
                    if creds and creds.expired and creds.refresh_token:
                        creds.refresh(Request())
                    print("Google Drive and Sheets connected successfully using OAuth Client (token.json).")
                except Exception as e:
                    print(f"OAuth Client authentication from token.json failed: {e}")
                    creds = None # Reset creds if token fails
            else:
                print("token.json not found. OAuth client authentication not available (interactive flow not supported in API directly).")

        # --- Build services if credentials obtained ---
        if creds:
            drive_service = build(
                "drive",
                "v3",
                credentials=creds,
                cache_discovery=False
            )
            sheets_service = build(
                "sheets",
                "v4",
                credentials=creds,
                cache_discovery=False
            )
        else:
            print("No valid Google credentials found after trying both Service Account and OAuth Client. Google API functionality will be limited.")


        # Assuming NGROK_PUBLIC_URL is provided by the user if ngrok is used
        # For Colab environments, if you are running a Django app and exposing it via ngrok
        # you would typically set this as an environment variable or retrieve it dynamically.
        # For demonstration purposes, you might hardcode it if it's stable during a session.
        # Example: NGROK_PUBLIC_URL = "https://your-ngrok-url.ngrok-free.app"
        NGROK_PUBLIC_URL = os.environ.get('NGROK_PUBLIC_URL', '') # User should set this env var if using ngrok
        # Hardcode your current ngrok URL here if it's easier for testing:
        # NGROK_PUBLIC_URL = "https://212390913ce5.ngrok-free.app" # Replace with your actual ngrok URL


        # ===========================
        # Generate public URLs (for local files, if used) and prepare response
        # ===========================
        if request.META.get('HTTP_HOST'):
            base_url = f"{request.scheme}://{request.META['HTTP_HOST']}"
        elif NGROK_PUBLIC_URL: # Use provided ngrok URL if available and not from request
            base_url = NGROK_PUBLIC_URL
        else:
            # This is a fallback/mock base_url. In a real deployed Django app,
            # request.META.get('HTTP_HOST') would provide the actual host.
            # For Colab, a local server or a public tunnel like ngrok would be needed to serve media.
            base_url = "http://localhost:8000" # Placeholder, generally not accessible externally in Colab
            print("Warning: base_url defaulted to localhost. For external access (e.g., from React frontend), consider setting NGROK_PUBLIC_URL or ensuring your Django app is publicly accessible.")


        local_file_url = f"{base_url}/media/{unique_folder_name}/{report_file_name}"
        html_preview_file_url = f"{base_url}/media/{unique_folder_name}/{html_preview_file_name}"

        # Determine final_download_url and final_preview_url based on local files if no Google Sheet is created
        final_download_url = local_file_url
        final_preview_url = local_file_url

        # ===========================
        # Create Google Sheet in specified folder
        # ===========================
        # Removed duplicate and incorrect Google Sheet creation logic from here.
        # The correct logic using drive_service.files().create is already present above.

        base_sheet_id=settings.MASTER_SHEET_ID # This is likely an existing sheet ID, not used for *creating* new one in a folder.
        timestamp_for_sheet=datetime.now().strftime("%Y%m%d%H%M%S")
        

        new_sheet_id = None
        google_sheet_link_for_response = None # Use a new variable to avoid confusion with the one initialized earlier

        if drive_service:
            try:
                # Use Drive API to create the spreadsheet directly in the specified folder
                file_metadata = file_metadata = {
            'name': f"Credit Card Reconciliation Report {timestamp_for_sheet}",
            'mimeType': 'application/vnd.google-apps.spreadsheet'
            # ❌ parents line REMOVE
        }

                created_spreadsheet = drive_service.files().create(
            body=file_metadata,
            fields='id'
        ).execute()
                new_sheet_id = created_spreadsheet.get('id')
                google_sheet_link_for_response = "https://docs.google.com/spreadsheets/d/{new_sheet_id}"
                print("New Google Sheet created in specified folder:", google_sheet_link_for_response)

                # Now that the sheet is created, if we need to write data to it, we'd use sheets_service.
                # This part is not in the error, but good to keep in mind.

            except HttpError as err:
                print(f"Error creating Google Sheet using Drive API (HttpError): {err}")

                google_sheet_link_for_response = None # Ensure it's None on error
            except Exception as e:
                print(f"Error creating Google Sheet in folder {FOLDER_ID} using Drive API (General Error): {e}")
                google_sheet_link_for_response = None # Ensure it's None if service not initialized
        else:
            print("Google Drive service is not initialized. Cannot create Google Sheet in a specific folder.")
            google_sheet_link_for_response = None # Ensure it's None if service not initialized

        # Update final_download_url and final_preview_url if a Google Sheet was successfully created
        if google_sheet_link_for_response:
            final_download_url = google_sheet_link_for_response
            final_preview_url = google_sheet_link_for_response

        if not bank.empty:
             try:
                min_dt = bank["DT"].min()
                max_dt = bank["DT"].max()
                if not pd.isna(min_dt) and not pd.isna(max_dt):
                    ReconciliationRecord.objects.create(
                        client_name=client_name,
                        min_date=min_dt.date(),
                        max_date=max_dt.date(),
                        total_transactions=len(bank),
                        bank_filename=bank_file_obj.name,
                        hotel_filename=hotel_file_obj.name,
                    )
             except Exception as e:
                 print(f"Failed to save reconciliation record: {e}")

        return Response({
            "status": "success",
            "success": True,
            "downloadUrl": final_download_url,
            "previewUrl": final_preview_url,
            "googleSheetLink": google_sheet_link_for_response, # Use the new variable here
            "reconciledCount": reconciledCount,
            "unreconciledCount": unreconciledCount,
            "totalEntries": totalEntries,
            "localFileUrl": local_file_url  # Optional
        })
    
